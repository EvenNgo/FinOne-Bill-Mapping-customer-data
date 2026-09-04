import io
import os
import re
import zipfile
import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="FinOne Data Mapper Pro", page_icon="📊", layout="wide"
)

# ==============================================================================
# 0. GIAO DIỆN HIỆN ĐẠI (CUSTOM CSS - FINTECH STYLE)
# ==============================================================================
st.markdown("""
<style>
    html, body, [class*="css"] {
        font-family: 'Inter', 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    .stApp { background: linear-gradient(180deg, #F6F9FC 0%, #EEF3F9 100%); }
    .heno-header {
        background: linear-gradient(120deg, #0B2A5B 0%, #114C9C 45%, #1785D6 100%);
        padding: 28px 36px; border-radius: 18px; margin-bottom: 28px;
        box-shadow: 0 10px 30px rgba(11, 42, 91, 0.25); color: white;
    }
    .heno-header h1 { margin: 0; font-size: 1.9rem; font-weight: 700; letter-spacing: -0.5px; color: #FFFFFF; }
    .heno-header p { margin: 6px 0 0 0; color: #D6E6FA; font-size: 0.95rem; }
    .heno-badge {
        display: inline-block; background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.3);
        padding: 4px 12px; border-radius: 999px; font-size: 0.75rem; color: #EAF3FF; margin-top: 12px; backdrop-filter: blur(4px);
    }
    .heno-card {
        background: #FFFFFF; border-radius: 16px; padding: 22px 26px; margin-bottom: 22px;
        border: 1px solid #E6ECF3; box-shadow: 0 4px 18px rgba(17, 76, 156, 0.06);
    }
    .heno-section-title { font-size: 1.05rem; font-weight: 700; color: #0B2A5B; margin-bottom: 4px; display: flex; align-items: center; gap: 8px; }
    .heno-section-sub { font-size: 0.85rem; color: #7A8AA0; margin-bottom: 16px; }
    .stTextInput input, .stNumberInput input, .stSelectbox > div > div, textarea {
        border-radius: 10px !important; border: 1px solid #DCE4EE !important; background-color: #FBFCFE !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus { border-color: #1785D6 !important; box-shadow: 0 0 0 3px rgba(23, 133, 214, 0.15) !important; }
    div[role="radiogroup"] label {
        background: #F5F8FC; border: 1px solid #E6ECF3; padding: 6px 14px; border-radius: 10px; margin-bottom: 6px !important; transition: all 0.15s ease;
    }
    div[role="radiogroup"] label:hover { border-color: #1785D6; background: #EAF3FD; }
    .stButton > button, .stDownloadButton > button { border-radius: 12px; font-weight: 600; padding: 0.6rem 1.4rem; border: none; transition: all 0.15s ease; }
    .stButton > button[kind="primary"] { background: linear-gradient(120deg, #114C9C 0%, #1785D6 100%); box-shadow: 0 6px 16px rgba(23, 133, 214, 0.35); }
    .stButton > button[kind="primary"]:hover { transform: translateY(-1px); box-shadow: 0 8px 20px rgba(23, 133, 214, 0.45); }
    .stDownloadButton > button { background: linear-gradient(120deg, #0F9D6A 0%, #12B77E 100%); color: white; box-shadow: 0 6px 16px rgba(15, 157, 106, 0.3); }
    .stDownloadButton > button:hover { transform: translateY(-1px); }
    [data-testid="stFileUploaderDropzone"] { background: #F5F9FF; border: 2px dashed #A9C7EA; border-radius: 14px; }
    .stProgress > div > div { background: linear-gradient(90deg, #114C9C, #1785D6); border-radius: 10px; }
    hr { border: none; border-top: 1px solid #E1E8F0; margin: 22px 0; }
    .stTabs [data-baseweb="tab"] { border-radius: 10px 10px 0 0; font-weight: 600; }
    [data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; border: 1px solid #E6ECF3; }
    div[data-testid="stAlert"] { border-radius: 12px; }
    .streamlit-expanderHeader { font-weight: 600 !important; color: #0B2A5B !important; background-color: #F8FAFC !important; border-radius: 8px !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="heno-header">
    <h1>📊 Chuyển Đổi Dữ Liệu Khách Hàng</h1>
    <p>Công cụ AI hỗ trợ ánh xạ dữ liệu khách hàng sang chuẩn nhập liệu FinOne Bill</p>
    <span class="heno-badge">FinOne Data Mapper Pro</span>
</div>
""", unsafe_allow_html=True)

# ==============================================================================
# 1. BỘ TỪ KHÓA MEGA & TỪ ĐIỂN (AI DATA BRAIN)
# ==============================================================================
KEYWORDS = {
    "stt": ["tt", "stt", "số tt", "số thứ tự", "no.", "no", "thứ tự"],
    "name": ["họ và tên", "họ tên", "tên khách hàng", "người đại diện", "chủ hộ", "học sinh", "tên học sinh", "người nộp", "bên a", "kh", "khách", "tên cháu", "tên hv", "học viên", "người mua", "tên đv", "tên", "đơn vị"],
    "last_name": ["họ đệm", "họ và đệm", "họ và tên đệm", "họ lót", "họ"],
    "first_name": ["tên gọi", "tên"],
    "phone": ["điện thoại", "sđt", "sdt", "phone", "tel", "di động", "mobile", "liên hệ", "đt cha", "đt mẹ", "hotline", "số đt"],
    "id": ["mã", "id", "định danh", "cccd", "cmnd", "cmt", "mã kh", "số định danh", "mã học sinh", "mã hv", "passport"],
    "group": ["khu vực", "nhóm", "phường", "xã", "tổ", "cụm", "khối", "lớp"]
}

COMMON_SURNAMES = {"nguyễn", "trần", "lê", "phạm", "hoàng", "huỳnh", "phan", "vũ", "võ", "đặng", "bùi", "đỗ", "hồ", "ngô", "dương", "lý", "đoàn", "chu", "trịnh", "đinh", "khất", "lâm", "thái", "phùng", "mai", "tô", "tôn", "đồng", "lục", "trương", "cáp"}
ORG_KEYWORDS = ["công ty", "cty", "doanh nghiệp", "trung tâm", "ubnd", "ủy ban", "phường", "quận", "huyện", "thành phố", "hợp tác xã", "htx", "trường", "ban chỉ huy", "ban quản lý", "chi nhánh", "cửa hàng", "đại lý", "nhà thuốc", "bệnh viện", "cơ sở", "tập đoàn", "sở", "bộ"]

# ==============================================================================
# 2. HÀM CORE & CACHE TỐC ĐỘ CAO
# ==============================================================================
@st.cache_data(show_spinner=False)
def sanitize_excel_bytes(file_bytes):
    try:
        in_zip = zipfile.ZipFile(io.BytesIO(file_bytes), "r")
        out_buf = io.BytesIO()
        out_zip = zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED)
        for item in in_zip.infolist():
            data = in_zip.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                data = re.sub(b"<(?:\w+:)?dataValidations[^>]*>.*?</(?:\w+:)?dataValidations>", b"", data, flags=re.DOTALL)
                data = re.sub(b"<(?:\w+:)?dataValidation[^>]*>.*?</(?:\w+:)?dataValidation>", b"", data, flags=re.DOTALL)
            out_zip.writestr(item, data)
        out_zip.close()
        out_buf.seek(0)
        return out_buf.getvalue()
    except Exception:
        return file_bytes

@st.cache_data(show_spinner=False)
def get_sheet_names(file_bytes):
    safe_bytes = sanitize_excel_bytes(file_bytes)
    return pd.ExcelFile(io.BytesIO(safe_bytes)).sheet_names

@st.cache_data(show_spinner=False)
def load_sheet(file_bytes, sheet_name, header=None, nrows=None):
    safe_bytes = sanitize_excel_bytes(file_bytes)
    return pd.read_excel(io.BytesIO(safe_bytes), sheet_name=sheet_name, header=header, nrows=nrows)

def auto_detect_header_row_smart(raw_df):
    best_row, max_matches = 0, 0
    for r_idx in range(min(15, len(raw_df))):
        row_vals = [str(v).lower().strip() for v in raw_df.iloc[r_idx].values if pd.notna(v)]
        matches = 0
        for kws in KEYWORDS.values():
            cat_matched = False
            for kw in kws:
                for cell_val in row_vals:
                    if len(kw) <= 3:
                        if kw == cell_val or kw in re.findall(r'\b\w+\b', cell_val):
                            cat_matched = True; break
                    else:
                        if kw in cell_val:
                            cat_matched = True; break
                if cat_matched: break
            if cat_matched: matches += 1
        if matches > max_matches:
            max_matches = matches
            best_row = r_idx
    if max_matches < 2: return 0, 0
    return best_row, max_matches

# ==============================================================================
# 3. ĐỘNG CƠ AI CẢM BIẾN DỮ LIỆU
# ==============================================================================
def resolve_column_super_sensor(df, target_col, category):
    cols = df.columns.tolist()
    if target_col and target_col not in ["-- Bỏ trống --", ">> Nhập giá trị cố định <<"]:
        for c in cols:
            if str(c).strip().lower() == str(target_col).strip().lower(): return c
                
    if category in KEYWORDS:
        for c in cols:
            c_str = str(c).lower().strip()
            for kw in KEYWORDS[category]:
                if len(kw) <= 3: 
                    if kw == c_str or kw in re.findall(r'\b\w+\b', c_str): return c
                else:
                    if kw in c_str: return c

    best_col, best_score = None, 0
    for c in cols:
        sample_raw = df[c].dropna().astype(str).str.strip().tolist()[:15]
        sample = [v for v in sample_raw if v.lower() not in ["none", "nan", "null", "", "0", "0.0"]]
        if len(sample) < 3: continue
        
        score = 0
        if category == "name":
            for val in sample:
                val_lower = val.lower()
                is_org = any(kw in val_lower for kw in ORG_KEYWORDS)
                
                if not is_org and re.search(r'\d', val): score -= 5
                if "@" in val or "http" in val: score -= 5 
                
                words = val.split()
                if is_org:
                    if 2 <= len(words) <= 15: score += 2 
                    score += 5 
                else:
                    if 2 <= len(words) <= 5: score += 2
                    if len(words) > 6: score -= 3 
                    if words and words[0].lower() in COMMON_SURNAMES: score += 5
                if val.istitle() or val.isupper(): score += 1 
                
        elif category == "last_name":
            for val in sample:
                if re.search(r'\d', val): score -= 5
                words = val.split()
                if 1 <= len(words) <= 3: score += 2
                if words and words[0].lower() in COMMON_SURNAMES: score += 5
                
        elif category == "first_name":
            for val in sample:
                if re.search(r'\d', val): score -= 5
                words = val.split()
                if len(words) == 1: score += 4
                if val.istitle(): score += 1
                
        elif category == "phone":
            for val in sample:
                val_no_space = val.replace(" ", "")
                if re.search(r'(?<!\d)0\d{9}(?!\d)', val_no_space): score += 5
                else: score -= 2
                
        elif category == "id":
            for val in sample:
                d = re.sub(r'\D', '', val.split(".")[0])
                if len(d) in [9, 11, 12]: score += 5
                else: score -= 2
                
        avg_score = score / len(sample)
        if avg_score > best_score and avg_score > 0.5:
            best_score = avg_score
            best_col = c
    return best_col

# ==============================================================================
# 4. CÁC HÀM LÀM SẠCH VÀ CHUẨN HÓA (ĐÃ FIX LỖI TÊN 1 CHỮ CÁI)
# ==============================================================================
def clean_phone(phone_val):
    if pd.isna(phone_val): return ""
    val_str = str(phone_val).split(".")[0].strip()
    if val_str.lower() in ["0", "none", "nan", "null", ""]: return ""
    val_str = re.sub(r'(?<!\d)\+?84(?=\d{8,9})', '0', val_str)
    val_no_space = val_str.replace(" ", "")
    matches = re.findall(r'(?<!\d)0\d{9}(?!\d)', val_no_space)
    if matches: return matches[0]
        
    cleaned = re.sub(r"[/,;\-–—|và]+", " ", val_str)
    tokens = cleaned.split()
    for tok in tokens:
        digits = re.sub(r"\D", "", tok)
        if len(digits) == 10 and digits.startswith("0"): return digits
            
    digits_all = re.sub(r"\D", "", val_str)
    if len(digits_all) == 10 and digits_all.startswith("0"): return digits_all
    return ""

def clean_id_val(raw_val):
    if pd.isna(raw_val): return ""
    val_str = str(raw_val).split(".")[0].strip()
    if val_str.lower() in ["none", "nan", "null", "", "0"]: return ""
    val_clean = re.sub(r"\s+", "", val_str)
    if len(val_clean) == 11 and val_clean.isdigit():
        val_clean = "0" + val_clean
    return val_clean

def is_valid_human_name(name):
    if not name: return False
    name_clean = str(name).strip()
    # CHO PHÉP TÊN 1 CHỮ CÁI (Như "Ý", "A") -> Sửa len < 2 thành len < 1 (chuỗi rỗng)
    if len(name_clean) < 1: return False
    
    if re.match(r"^[\d\s\.\,\-\_]+$", name_clean): return False
    if name_clean.lower() in ["họ và tên", "họ tên", "tên", "người liên hệ", "tổng cộng", "stt", "nan", "none"]: return False
    return True

# ==============================================================================
# 5. XỬ LÝ LÕI TỪNG SHEET VÀ THEO DÕI TỌA ĐỘ DÒNG EXCEL
# ==============================================================================
def process_sheet_data(s_name, file_bytes, config, file_name=""):
    raw_df = load_sheet(file_bytes, sheet_name=s_name, header=None, nrows=15)
    if raw_df.empty: return [], [], "Trống", True

    auto_h, matches = auto_detect_header_row_smart(raw_df)
    
    # Xác định dòng offset để tính chuẩn xác Dòng Excel
    if matches == 0 and not config["apply_fixed_header"]:
        df = load_sheet(file_bytes, sheet_name=s_name, header=None).dropna(how="all")
        df.columns = [f"Col_{i}" for i in range(len(df.columns))]
        hdr_text = "AI Cảm Biến tự dò (Không Header)"
        header_offset = -1 # Không có tiêu đề
    else:
        use_h_idx = config["header_row_idx"] if config["apply_fixed_header"] else auto_h
        df = load_sheet(file_bytes, sheet_name=s_name, header=use_h_idx).dropna(how="all")
        hdr_text = f"Dòng {use_h_idx + 1}"
        header_offset = use_h_idx

    s_name_col = resolve_column_super_sensor(df, config["map_name"] if config["name_mode"] == "Họ và Tên gộp chung 1 cột" else None, "name")
    s_last_col = resolve_column_super_sensor(df, config["map_last_name"], "last_name")
    s_first_col = resolve_column_super_sensor(df, config["map_first_name"], "first_name")
    s_phone_col = resolve_column_super_sensor(df, config["map_phone"], "phone")
    s_id_col = resolve_column_super_sensor(df, config["map_id"], "id")

    if config["name_mode"] == "Họ và Tên gộp chung 1 cột" and not s_name_col:
        s_last_col = resolve_column_super_sensor(df, None, "last_name")
        s_first_col = resolve_column_super_sensor(df, None, "first_name")
        
    is_missing_name = False
    if not s_name_col and not (s_last_col and s_first_col):
        is_missing_name = True

    if is_missing_name: 
        return [], [], hdr_text, True

    v_rows, r_rows = [], []
    for row_idx_pandas, r in df.iterrows():
        # CÔNG THỨC TÍNH CHUẨN XÁC DÒNG EXCEL (1-indexed)
        # r.name là index gốc của pandas (kể cả khi đã dropna các dòng trống ở giữa)
        excel_row_num = header_offset + 2 + r.name 
        
        if s_name_col:
            full_name = str(r.get(s_name_col, "")).strip()
        else:
            p_last = str(r.get(s_last_col, "")).strip() if pd.notna(r.get(s_last_col)) else ""
            p_first = str(r.get(s_first_col, "")).strip() if pd.notna(r.get(s_first_col)) else ""
            if p_last.lower() in ['nan', 'none']: p_last = ""
            if p_first.lower() in ['nan', 'none']: p_first = ""
            full_name = re.sub(r"\s+", " ", f"{p_last} {p_first}").strip()

        if not is_valid_human_name(full_name):
            r_rows.append({
                "File Nguồn": file_name, 
                "Sheet": s_name, 
                "Dòng Excel": f"Dòng {excel_row_num}", # Đưa tọa độ vào file báo lỗi
                "Dữ liệu Tên": full_name, 
                "Lý do Loại": "Tên rỗng hoặc chứa số/ký tự sai định dạng"
            })
            continue
            
        phone_raw = config["fix_phone"] if config["map_phone"] == ">> Nhập giá trị cố định <<" else r.get(s_phone_col, "")
        id_raw = config["fix_id"] if config["map_id"] == ">> Nhập giá trị cố định <<" else r.get(s_id_col, "")
        
        if "Tên từng File" in config["group_strategy"]: grp_val = file_name
        elif "Tên từng Sheet" in config["group_strategy"]: grp_val = s_name
        elif "Một cột trong bảng" in config["group_strategy"]:
            grp_col = resolve_column_super_sensor(df, config["map_group"], "group")
            grp_val = str(r.get(grp_col, config["fix_group"])) if grp_col and pd.notna(r.get(grp_col)) else config["fix_group"]
        else:
            grp_val = config["fix_group"]
            
        if str(grp_val).lower() in ['nan', 'none', '']: grp_val = config["fix_group"]

        v_rows.append({
            "Cơ sở (*)": config["val_coso"],
            "Nhóm KH (*)": grp_val,
            "Tên KH (*)": full_name,
            "Mã định danh": clean_id_val(id_raw),
            "Điện thoại (*)": clean_phone(phone_raw),
            "Email": "", "Loại KH": "", "Địa chỉ/Ghi chú": "", "Tên doanh nghiệp": "",
        })
        
    return v_rows, r_rows, hdr_text, False

# ==============================================================================
# 6. GIAO DIỆN STREAMLIT
# ==============================================================================
st.markdown('<div class="heno-card">', unsafe_allow_html=True)
st.markdown('<div class="heno-section-title">📁 Tải Dữ Liệu Lên</div>', unsafe_allow_html=True)
st.markdown('<div class="heno-section-sub">Hỗ trợ nhiều file Excel cùng lúc, hệ thống sẽ tự động dò cột phù hợp</div>', unsafe_allow_html=True)
uploaded_files = st.file_uploader("1. Tải lên 1 hoặc nhiều file Excel (.xlsx, .xls):", type=["xlsx", "xls"], accept_multiple_files=True)
st.markdown('</div>', unsafe_allow_html=True)

if uploaded_files:
    file_map = {f.name: f.getvalue() for f in uploaded_files}
    st.success(f"📂 Đã nạp thành công **{len(uploaded_files)} file**")

    st.markdown('<div class="heno-card">', unsafe_allow_html=True)
    col_f1, col_f2 = st.columns([1, 2])
    with col_f1: sample_file_name = st.selectbox("🎯 Chọn File mẫu:", options=list(file_map.keys()))
    sample_file_bytes = file_map[sample_file_name]
    sample_sheets = get_sheet_names(sample_file_bytes)

    default_sheets = [s for s in sample_sheets if s.lower() not in ["dulieu", "thongbaoloi"]] or sample_sheets
    with col_f2: selected_sheets = st.multiselect("📋 Danh sách Sheet xử lý:", options=sample_sheets, default=default_sheets)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="heno-card">', unsafe_allow_html=True)
    st.markdown('<div class="heno-section-title">🔗 BƯỚC 1: KIỂM SOÁT TIÊU ĐỀ & GHÉP CỘT (AI HỖ TRỢ)</div>', unsafe_allow_html=True)

    preview_ref_sheet = selected_sheets[0] if selected_sheets else sample_sheets[0]
    raw_preview = load_sheet(sample_file_bytes, sheet_name=preview_ref_sheet, header=None, nrows=15)
    detected_h_row, _ = auto_detect_header_row_smart(raw_preview)

    c_h1, c_h2 = st.columns([1, 2])
    with c_h1: header_choice = st.number_input("📌 Vị trí dòng tiêu đề (Header):", min_value=1, max_value=15, value=int(detected_h_row + 1))
    with c_h2: apply_fixed_header = st.checkbox("Khóa cứng dòng tiêu đề này cho mọi file (Bỏ tích để AI tự do quét từng file)", value=False)

    curr_header_idx = int(header_choice) - 1
    sample_cols_df = load_sheet(sample_file_bytes, sheet_name=preview_ref_sheet, header=curr_header_idx, nrows=5)
    valid_cols = [str(c).strip() for c in sample_cols_df.columns if not str(c).startswith("Unnamed:") and pd.notna(c)]
    dropdown_opts = ["-- Bỏ trống --", ">> Nhập giá trị cố định <<"] + valid_cols

    val_coso = st.text_input("🏢 Tên Cơ sở (Áp dụng tất cả):", placeholder="Ví dụ: Cơ sở A...")

    col_map1, col_map2 = st.columns(2)
    with col_map1:
        grp_strategy = st.radio("🏢 Nhóm Khách hàng:", ["Tên từng File", "Tên từng Sheet", "Nhập tên cố định", "Một cột trong bảng"], index=0)
        f_group, m_group = "", None
        if grp_strategy == "Nhập tên cố định": f_group = st.text_input("✍️ Nhập Nhóm cố định:")
        elif grp_strategy == "Một cột trong bảng": 
            m_group = st.selectbox("Cột Nhóm:", dropdown_opts)
            f_group = st.text_input("Dự phòng nếu ô trống:")

        name_mode = st.radio("👤 Định dạng cột Họ Tên:", ["Tách riêng 2 cột (Họ đệm + Tên)", "Họ và Tên gộp chung 1 cột"], index=1)
        m_name, m_last, m_first = None, None, None
        if name_mode == "Họ và Tên gộp chung 1 cột": m_name = st.selectbox("Cột Họ và Tên:", dropdown_opts)
        else:
            m_last = st.selectbox("Cột Họ đệm:", dropdown_opts)
            m_first = st.selectbox("Cột Tên:", dropdown_opts)

    with col_map2:
        m_phone = st.selectbox("Cột Điện thoại:", dropdown_opts)
        f_phone = st.text_input("✍️ SĐT cố định:") if m_phone == ">> Nhập giá trị cố định <<" else ""
        m_id = st.selectbox("Cột Mã định danh:", dropdown_opts)
        f_id = st.text_input("✍️ Mã cố định:") if m_id == ">> Nhập giá trị cố định <<" else ""
        
        st.info("🧠 **Sensor Đang Hoạt Động:** Nếu các cột bạn chọn ở trên không tồn tại ở các file khác, hệ thống sẽ tự động quét Data (chữ, số) để tìm và ghép cột Tên & Điện thoại chính xác.")
    st.markdown('</div>', unsafe_allow_html=True)

    config = {
        "val_coso": val_coso.strip(), "header_row_idx": curr_header_idx, "apply_fixed_header": apply_fixed_header,
        "group_strategy": grp_strategy, "fix_group": f_group, "map_group": m_group,
        "name_mode": name_mode, "map_name": m_name, "map_last_name": m_last, "map_first_name": m_first,
        "map_phone": m_phone, "fix_phone": f_phone, "map_id": m_id, "fix_id": f_id,
    }

    # ==============================================================================
    # BƯỚC 2: XUẤT FILE HOÀN CHỈNH & ĐỐI SOÁT BÁO LỖI
    # ==============================================================================
    st.markdown('<div class="heno-card">', unsafe_allow_html=True)
    st.markdown('<div class="heno-section-title">🚀 BƯỚC 2: XÁC NHẬN VÀ XUẤT TOÀN BỘ FILE</div>', unsafe_allow_html=True)
    output_filename = st.text_input("Tên file xuất ra:", value="Ket_Qua_Nhap_Lieu_FinOne.xlsx")

    if st.button("🚀 GỘP VÀ XUẤT FILE FINONE", type="primary"):
        if not val_coso.strip():
            st.error("🚨 LỖI: Bắt buộc điền 'Tên Cơ sở'.")
            st.stop()

        wb = openpyxl.load_workbook("mau-nhap-lieu-khach-hang.xlsx")
        ws = wb["Bảng nhập liệu khách hàng"] if "Bảng nhập liệu khách hàng" in wb.sheetnames else wb.active

        total_valid, total_rejected = 0, 0
        summary_stats, all_rejected_rows, bulk_write_data = [], [], []
        sheets_missing_cols = []

        progress_bar = st.progress(0)
        
        for f_idx, (fname, fbytes) in enumerate(file_map.items()):
            base_n = os.path.splitext(fname)[0]
            target_sheets = [s for s in get_sheet_names(fbytes) if s.lower() not in ["dulieu", "thongbaoloi"]]
            if not target_sheets: target_sheets = get_sheet_names(fbytes)

            for s_name in target_sheets:
                v_rows, r_rows, hdr_text, is_missing = process_sheet_data(s_name, fbytes, config, base_n)

                if is_missing: sheets_missing_cols.append(f"{fname} - {s_name}")
                
                bulk_write_data.extend(v_rows)
                total_valid += len(v_rows)
                
                all_rejected_rows.extend(r_rows)
                total_rejected += len(r_rows)

                summary_stats.append({
                    "Tên File": fname, "Tên Sheet": s_name, "Dòng Header": hdr_text,
                    "Số dòng Hợp lệ": len(v_rows), "Số dòng Bị loại": len(r_rows),
                })
            progress_bar.progress((f_idx + 1) / len(file_map))

        col_mapping = {"Cơ sở (*)": 2, "Nhóm KH (*)": 3, "Tên KH (*)": 4, "Mã định danh": 5, "Điện thoại (*)": 6, "Email": 7, "Loại KH": 8, "Địa chỉ/Ghi chú": 9, "Tên doanh nghiệp": 10}
        
        current_row = 6
        for row_data in bulk_write_data:
            for key, col_idx in col_mapping.items():
                cell_val = row_data.get(key, "")
                safe_val = "" if str(cell_val).strip().lower() in ["", "none", "nan", "null", "0", "0.0"] or pd.isna(cell_val) else str(cell_val).strip()
                
                cell = ws.cell(row=current_row, column=col_idx)
                
                if key in ["Điện thoại (*)", "Mã định danh"]:
                    cell.data_type = 's'  
                    cell.number_format = '@'  
                    cell.value = safe_val
                else:
                    cell.value = safe_val
            current_row += 1

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        st.success(f"🎉 Đã gộp thành công **{total_valid} khách hàng**!")

        if sheets_missing_cols:
            st.error("🚨 **CẢNH BÁO: AI Cảm biến KHÔNG THỂ tìm thấy Tên khách hàng ở các sheet sau (Dữ liệu bị bỏ qua):**\n\n" + "\n".join(f"- {s}" for s in sheets_missing_cols))

        st.info(f"✅ **Đối soát quân số:** Tổng dòng quét = **{total_valid + total_rejected}** | Hợp lệ = **{total_valid}** | Bị loại = **{total_rejected}**")

        tab_final1, tab_final2 = st.tabs(["📥 Tải File Hoàn Chỉnh (FinOne)", "❌ Báo Cáo Tổng Hợp Dòng Bị Loại"])

        with tab_final1:
            st.dataframe(pd.DataFrame(summary_stats), use_container_width=True)
            st.download_button("📥 TẢI FILE EXCEL KẾT QUẢ", output_buffer, output_filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        with tab_final2:
            if all_rejected_rows:
                df_all_rej = pd.DataFrame(all_rejected_rows)
                
                st.markdown('<div class="heno-section-sub" style="margin-bottom: 10px;">👇 Nhấn vào từng File/Sheet bên dưới để xem chi tiết các dữ liệu bị từ chối:</div>', unsafe_allow_html=True)
                
                error_files = df_all_rej['File Nguồn'].unique()
                for f_err in error_files:
                    df_file = df_all_rej[df_all_rej['File Nguồn'] == f_err]
                    error_sheets = df_file['Sheet'].unique()
                    
                    for s_err in error_sheets:
                        df_sheet = df_file[df_file['Sheet'] == s_err]
                        with st.expander(f"📁 {f_err}  👉  📄 Sheet: {s_err}  |  ❌ Từ chối {len(df_sheet)} dòng"):
                            # HIỂN THỊ CỘT DÒNG EXCEL RA BẢNG CHI TIẾT
                            st.dataframe(df_sheet[['Dòng Excel', 'Dữ liệu Tên', 'Lý do Loại']], use_container_width=True)
                
                st.markdown("---")
                rej_buf = io.BytesIO()
                df_all_rej.to_excel(rej_buf, index=False, sheet_name="Dong_Bi_Loai", engine="openpyxl")
                rej_buf.seek(0)
                st.download_button("📥 TẢI FILE EXCEL DANH SÁCH LỖI MỞ RỘNG", rej_buf, "Tong_Hop_Dong_Bi_Loai.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.success("🎉 Toàn bộ dữ liệu đều hợp lệ 100%!")
    st.markdown('</div>', unsafe_allow_html=True)
